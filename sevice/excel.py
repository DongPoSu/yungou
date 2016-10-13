# coding=utf8

from openpyxl import Workbook
from openpyxl import load_workbook

from common.DealStatus import PAY_SUCCESS, PAY_FAILED
from constants.db_constants import BILL_DEAL, NORMAL_DEAL
from dao import Deal
from sevice.member_deal import query_deal, check_bank_deal


# 导出银行卡审核通过提现记录
def export_bank_deal(check_start_date, check_end_date, title):
    '''
    :param check_start_date: 审核开始日期
    :param check_end_date:  审核结束日期
    :param title: 文件名
    :return:
    '''
    deals = query_deal(check_start_date, check_end_date, BILL_DEAL)
    wb = Workbook()
    ws = wb.active
    head_line = ['提现号', '审核日期', '序号', '币种', '金额', '收款人账号', '收款人名称', '收款账号开户行名称', '收款省份/收款银行',
                 '收款地市', '地区代码', '付款账号开户行名称', '付款人账号/卡号', '付款人名称 / 卡名称',
                 '汇款用途', '备注', '预约付款日期', '汇款方式', '收款账户短信通知手机号码', '自定义序号', '协议编号']
    ws.append(head_line)
    count = 0
    for deal in deals:
        count += 1
        ws.append(
            [deal.deal_code, deal.check_date.strftime('%Y-%m-%d %H:%M:%S'), count, "人民币", round(float(deal.apply_money) * 0.99,2),
             deal.bank_account, deal.bank_user, deal.bankcard_address,
             deal.bankcard_province,
             deal.bankcard_city, deal.bank_id, "工行广州花都雅居乐支行", "3602202119100259501", "广州思埠网络开发有限公司", "代付款", "",
             title, "加急", deal.phone, "", ""])

    wb.save("resources/%s银行卡打款名单.xlsx" % (title))


# 导出微信提现打款结果
def export_wx_deal(give_start_date, give_end_date, title):
    '''
    :param give_start_date: 打款开始时间
    :param give_end_date: 打款结束时间
    :param title:
    :return:
    '''
    deals = query_deal(give_start_date, give_end_date, NORMAL_DEAL)
    wb = Workbook()
    ws = wb.active
    head_line = ['序号', '姓名', '手机号', '提现金额', '提现状态', '申请日期', '打款备注', '打款日期']
    ws.append(head_line)
    count = 0
    for deal in deals:
        count += 1
        ws.append(
            [count, deal.user_name, deal.phone, deal.apply_money, deal.deal_status,
             deal.apply_date.strftime('%Y-%m-%d %H:%M:%S'), deal.give_invoice,
             deal.give_date.strftime('%Y-%m-%d %H:%M:%S')])
    wb.save("resources/%s微信打款结果.xlsx" % (title))


def import_order_trade_id(filename):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb['Sheet1']  # ws is now an IterableWorksheet
    list = []
    for row in ws.rows:
        if str(row[0].value) == "":
            pass
        data = str(row[0].value), str(row[1].value)
        list.append(data)


def import_bank_deal(filename):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb["Sheet1"]
    list = []
    count = 0
    for row in ws.rows:
        if count is 0:
            count += 1
            continue

        deal_result = Deal.BillDealResult()
        deal_result.bank_account = str(row[2].value).strip("\t").strip(" ")
        deal_result.bank_user=row[3].value
        deal_result.apply_money = int(row[6].value * 100)
        if str(row[8].value) == '处理成功':
            deal_result.deal_status = PAY_SUCCESS
        else:
            deal_result.deal_status = PAY_FAILED
        if row[9].value is None:
            deal_result.give_invoice = ""
        else:
            deal_result.give_invoice = row[9].value
        list.append(deal_result)
    check_bank_deal(list)





