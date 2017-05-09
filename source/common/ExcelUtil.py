from openpyxl import Workbook


class WriteExcel:
    _wb = Workbook()
    _file_path = "resources/"
    _file_name = ""
    _sheet_name = "Sheet1"
    _sheet_index = 0
    _ws = None

    def __init__(self, file_name):
        self._file_name = file_name

    def set_save_path(self, file_path):
        self._file_path = file_path

    def create_sheet(self, sheet_name=None, sheet_index=None):
        if sheet_name != "" and sheet_name is not None:
            self._sheet_name = sheet_name
        if sheet_index is not None and isinstance(sheet_index, int):
            self._sheet_index = sheet_index
        self._ws = self._wb.create_sheet(self._sheet_name, self._sheet_index)

    def write_content(self, data):
        if data is not None:
            self._ws.append(data)

    def write_close(self):
        self._wb.save(self._file_path + self._file_name + ".xlsx")
